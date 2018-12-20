# -*- coding=utf-8 -*-
import os, datetime, re, time, random, logging, traceback, pickle
from openpyxl import Workbook
from openpyxl import load_workbook
from zhihu_oauth import ZhihuClient, Answer, Article
from zhihu_oauth.exception import NeedCaptchaException

# 知乎经典19764022
# 破万赞29962923
# 妹子语言20121617
# 奸商技巧19659999
# 知识储备19598423
# 多看几遍19561986
# 醍醐灌顶19649021
# 淘宝宝贝20801936
# 神回复20186209
# 高质量实物19770547
# 视野19596498
# 搞笑的事19863027
# 相处28294827
collection_id = input('请输入收藏夹id号，获取所有答案：')
file_name = input('请输入保存文件名，不带后缀：')
TOKEN_FILE = 'token.pkl'
client = ZhihuClient()

def get_answers(answer,author):
    item_data = [datetime.datetime.now().strftime('%Y-%m-%d-%H:%M:%S'),re.compile(r'<.*?>', re.S).sub('', answer.content),
                answer.question.title, author]
    try:
        gender_dict = {'0': '女', '1': '男', '-1': '不详'}
        if answer.author.gender != None:
            gender = gender_dict[str(answer.author.gender)]
        else:
            gender = '不详'
        loc = ''
        if answer.author.locations:
            for location in answer.author.locations:
                loc += location.name
        company = ''
        job = ''
        if answer.author.employments:
            for employment in answer.author.employments:
                if 'company' in employment:
                    company += employment.company.name
                if 'job' in employment:
                    job += employment.job.name
        if answer.author.business:
            business = answer.author.business.name
        else:
            business = ''
        item_data += [gender, loc, business, company, job, datetime.datetime.fromtimestamp(answer.created_time),
                datetime.datetime.fromtimestamp(answer.updated_time), answer.voteup_count, answer.comment_count, answer.thanks_count
                ]
        sheet.append(item_data)
    except:
        sheet.append(item_data)
        wb.save('知乎-{}.xlsx'.format(file_name))
        # queue.remove(author)
        pickle.dump(queue, open("queue-collec.pkl","wb"))
        logging.error(answer.question.title+'******'+author+'\n'+'-'*60+'\n'+traceback.format_exc()+'-'*60+'\n')
    finally:
        pass
        # answer.save(answer.question.title)
        


if os.path.isfile(TOKEN_FILE):
    client.load_token(TOKEN_FILE)
else:
    try:
        client.login('email_or_phone', 'password')
    except NeedCaptchaException:
        with open('a.gif', 'wb') as f:
            f.write(client.get_captcha())
        captcha = input('please input captcha:')
        client.login('email_or_phone', 'password', captcha)
    client.save_token(TOKEN_FILE)
collection = client.collection(int(collection_id))
print(collection.title)
#日志设置
logging.basicConfig(level=logging.ERROR,  
                format='%(asctime)s %(levelname)s %(message)s',  
                datefmt='%Y-%m-%d %H:%M:%S',
                filename='zhi.log',
                filemode='w')

if os.path.exists('知乎-{}.xlsx'.format(file_name)):
    queue = pickle.load(open("queue-collec.pkl", "rb"))
    wb = load_workbook('知乎-{}.xlsx'.format(file_name))
    sheet = wb.active
    data_rows = sheet.max_row - 1
    print("上次进度已加载！")
else:
    queue = set()
    wb = Workbook()
    sheet = wb.active
    sheet.title = "知乎"
    item_name = ['time_now', 'content', 'que_title', 'author', 'gender', 'loc', 'business', 'company', 'job', 'created_time', 'updated_time', 'voteup_count', 'comment_count', 'thanks_count']
    for j,title in enumerate(item_name):
        sheet.cell(row=1, column=j+1).value = title
for num,content in enumerate(collection.contents):
    try:
        author = content.author.name
        if author not in queue and isinstance(content, Answer):
            queue.add(author)
            print(content.question.title)
            get_answers(content,author)
    except:
        wb.save('知乎-{}.xlsx'.format(file_name))
        pickle.dump(queue, open("queue-collec.pkl","wb"))
        logging.error(content.question.title+'******'+author+'\n'+'-'*60+'\n'+traceback.format_exc()+'-'*60+'\n')
    # elif isinstance(content, Article):
    #     article = content
    #     print(article.title)
    # if (num+1) % 20 == 0:
    #     wb.save('知乎-{}.xlsx'.format(file_name))
