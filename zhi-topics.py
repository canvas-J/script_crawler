# -*- coding=utf-8 -*-
import os, datetime, re, time, random, logging, traceback, pickle, hashlib
from openpyxl import Workbook
from openpyxl import load_workbook
from zhihu_oauth import ZhihuClient
from zhihu_oauth.exception import NeedCaptchaException


topic_id = input('请输入话题id号，获取所有答案：')
file_name = input('请输入保存文件名，不带后缀：')
TOKEN_FILE = 'token.pkl'
client = ZhihuClient()

if os.path.isfile(TOKEN_FILE):
    client.load_token(TOKEN_FILE)
else:
    try:
        client.login('email_or_phone', 'password')
    except NeedCaptchaException:
        with open('a.gif', 'wb') as f:
            f.write(client.get_captcha())
        captcha = input('please input captcha:')
        client.login('email_or_phone', 'password', captcha)
    client.save_token(TOKEN_FILE)

topic = client.topic(int(topic_id))
print(topic.name)

#日志设置
logging.basicConfig(level=logging.ERROR,  
                format='%(asctime)s %(levelname)s %(message)s',  
                datefmt='%Y-%m-%d %H:%M:%S',
                filename='zhi.log',
                filemode='w')

if os.path.exists('知乎-{}.xlsx'.format(file_name)):
    queue = pickle.load(open("queue.pkl", "rb"))
    wb = load_workbook('知乎-{}.xlsx'.format(file_name))
    sheet = wb.active
    data_rows = sheet.max_row - 1
    print("上次进度已加载！")
else:
    queue = set()
    wb = Workbook()
    sheet = wb.active
    sheet.title = "知乎"
    item_name = ['time_now', 'content', 'que_title', 'author', 'gender', 'loc', 'business', 'company', 'job', 'created_time', 'updated_time', 'voteup_count', 'comment_count', 'thanks_count']
    for j,title in enumerate(item_name):
        sheet.cell(row=1, column=j+1).value = title

num = 0
for question in topic.unanswered_questions:
    if question.title not in queue:
        print(question.title)
        hash_info = hashlib.md5(question.title.encode("utf-8")).hexdigest()
        queue.add(hash_info)
        for answer in question.answers:
            num += 1
            try:
                gender_dict = {'0': '女', '1': '男', '-1': '不详'}
                if answer.author.gender != None:
                    gender = gender_dict[str(answer.author.gender)]
                else:
                    gender = '不详'
                loc = ''
                if answer.author.locations:
                    for location in answer.author.locations:
                        loc += location.name
                company = ''
                job = ''
                if answer.author.employments:
                    for employment in answer.author.employments:
                        if 'company' in employment:
                            company += employment.company.name
                        if 'job' in employment:
                            job += employment.job.name
                if answer.author.business:
                    business = answer.author.business.name
                else:
                    business = ''
                item_data = [datetime.datetime.now().strftime('%Y-%m-%d-%H:%M:%S')]
                item_data += [re.compile(r'<.*?>', re.S).sub('', answer.content), question.title, answer.author.name, gender, loc, business,
                        company, job, datetime.datetime.fromtimestamp(answer.created_time),
                        datetime.datetime.fromtimestamp(answer.updated_time), answer.voteup_count, answer.comment_count, answer.thanks_count
                        ]
            except:
                wb.save('知乎-{}.xlsx'.format(file_name))
                queue.remove(hash_info)
                pickle.dump(queue, open("queue.pkl", "wb"))
                logging.error(question.title+'******'+answer.author.name+'\n'+'-'*60+'\n'+traceback.format_exc()+'-'*60+'\n')
            finally:
                sheet.append(item_data)
                # answer.save(question.title)
            if num % 20 == 0:
                wb.save('知乎-{}.xlsx'.format(file_name))